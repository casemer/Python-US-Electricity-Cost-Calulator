import openpyxl
from openpyxl import load_workbook

#Introduction

print(
    '------------------------------------------------\n'
    '            ELECTRICITY CALCULATOR              \n'
    '------------------------------------------------\n'
    'Instructions: Enter the name of the device then \n'
    'a space followed by the average number of hours \n'
    'you use the appliances (this can be a decimal). \n'
    'Use another space, repeat this process for each \n'
    'appliance you wish to use. Then hit "enter" for \n'
    'the next step of entering your lights. How many \n'
    'how long they are on, and their average wattage \n'
    'Finally, put in your state abv. or US for avg.  \n'
    )

app = input('Appliances: ')
light = input('Lights (how many, how long, and watts): ')
state = input('State (abv.): ')
split_light = light.split()
split_app = app.split()

kWh = 0;
lightkWh = 0;
eCost = 0;

#Splits up the appliances, checks each even entry which
#are the words, and searches for them on the excel sheet
for i in range(len(split_app)):
    watt = 0;
    hours = 0;
    if i % 2 == 0:

        def FindXlCell(search_str, range=None):
            global ws
            
            if not range:
                range = ws.iter_rows() # Defaults to whole sheet

            for tupleOfCells in range:
                for cell in tupleOfCells:
                    if (cell.value == search_str):
                        return [_tuple[0] for _tuple in ws.iter_cols(min_row=cell.row, max_row=cell.row)]

        wb = openpyxl.load_workbook("kWh.xlsx")
        ws = wb.worksheets[0]
        search_str = split_app[i].lower()
        #looks for the searched string which is each appliance
        cellsOfFoundRow = FindXlCell(search_str, ws.iter_rows(min_col=0, max_col=0))

        if cellsOfFoundRow:
            watt = str(cellsOfFoundRow[2].value)
        else:
            print("Could not find '{}' in the given cell range!".format(search_str))

        
        hours = split_app[i+1]

#Math for lights and their wattage
    kWh += ((float(watt) * float(hours))/1000)
    watt = split_light[2]
    lightkWh = (((int(watt) * int(split_light[1]))/1000) * int(split_light[0])) 


#Gets the state you put in and the cost of kWh on average there
search_str = state.upper()
cellsOfFoundRow = FindXlCell(search_str, ws.iter_rows(min_col=5, max_col=5))

if cellsOfFoundRow:
    eCost = str(cellsOfFoundRow[5].value)
else:
    print("Could not find '{}' in the given cell range!".format(search_str))


#final math getting the combined appliacnes and lights.
#then multiplying it by the state avg
cost = (kWh + lightkWh) * float(eCost)

print('Appliance Cost Per Day: $' + str(round(cost, 2)))
print('Appliance Cost Per Week: $' + str(round(cost * 7, 2)))
print('Appliance Cost Per Month: $' + str(round(cost * 30, 2)))
print('Appliance Cost Per Year: $' + str(round(cost * 365, 2)))



answer = input('Start Again? (y/n): ')
if answer == 'y':
    exec(open('kwhexcel.py').read())
else:
    print('Okay, Goodbye!')
    exit()
